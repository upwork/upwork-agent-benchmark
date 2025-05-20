"""Tools for handling project outputs."""

import csv
import json
import os
from datetime import datetime
from typing import Dict, List, Union

import docx
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


def write_to_docx(input_text: str, output_path: str) -> None:
  """Writes a string to a docx file at the specified output path.

  Args:
      input_text (str): The text content to write to the docx file
      output_path (str): The path where the docx file will be saved

  Returns:
      None
  """
  # Create a new Document
  doc = docx.Document()

  # Add the input text to the document
  doc.add_paragraph(input_text)

  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Save the document
  doc.save(output_path)

  return f'Successfully wrote text to {output_path}'


def write_to_xlsx(data: Union[List[List], Dict[str, List[List]]], output_path: str) -> str:
  """Writes data to an Excel file at the specified output path.

  Args:
      data: Either a 2D list of data for a single sheet, or a dictionary
           where keys are sheet names and values are 2D lists of data
      output_path (str): The path where the Excel file will be saved

  Returns:
      str: A success message with the output path
  """
  # Create a new workbook
  wb = Workbook()

  # Remove default sheet
  default_sheet = wb.active
  wb.remove(default_sheet)

  # If data is a dictionary, create multiple sheets
  if isinstance(data, dict):
    for sheet_name, sheet_data in data.items():
      # Create a new sheet
      ws = wb.create_sheet(title=sheet_name)

      # Add data to the sheet
      for row_idx, row in enumerate(sheet_data, 1):
        for col_idx, value in enumerate(row, 1):
          ws.cell(row=row_idx, column=col_idx, value=value)

  # If data is a list, create a single sheet
  else:
    # Create a sheet named "Sheet1"
    ws = wb.create_sheet(title='Sheet1')

    # Add data to the sheet
    for row_idx, row in enumerate(data, 1):
      for col_idx, value in enumerate(row, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Save the workbook
  wb.save(output_path)

  return f'Successfully wrote data to Excel file at {output_path}'


def write_to_json(data: dict, output_path: str) -> str:
  """Writes a dictionary to a JSON file at the specified output path.

  Args:
      data (dict): The dictionary to write to the JSON file
      output_path (str): The path where the JSON file will be saved

  Returns:
      str: A success message with the output path
  """
  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Write the dictionary to a JSON file
  with open(output_path, 'w') as f:
    json.dump(data, f, indent=2)

  return f'Successfully wrote JSON data to {output_path}'


def write_to_code_file(input_text: str, output_path: str, language: str = 'python') -> None:
  """Writes code to a file at the specified output path.

  Args:
      input_text (str): The code content to write to the file
      output_path (str): The path where the file will be saved
      language (str): Programming language for proper formatting
  Returns:
      None
  """
  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Format the code with headers
  formatted_code = format_code(input_text, language)

  # Write the code to file
  with open(output_path, 'w') as f:
    f.write(formatted_code)


def format_code(code: str, language: str) -> str:
  """Format code with language-specific headers and documentation."""
  # Get comment style for the language
  comment = {
    'python': '#',
    'javascript': '//',
    'typescript': '//',
    'java': '//',
    'c': '//',
    'cpp': '//',
    'ruby': '#',
    'php': '//',
    'go': '//',
    'rust': '//',
  }.get(language.lower(), '#')

  # Create header
  header = []

  # Add shebang for Python
  if language.lower() == 'python':
    header.append('#!/usr/bin/env python3')

  # Add language and timestamp
  header.append(f'{comment} Language: {language}')
  header.append(f'{comment} Created: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')
  header.append('')  # Blank line after header

  # Combine header and code
  return '\n'.join(header) + '\n' + code.strip()


def write_to_pdf(input_text: str, output_path: str) -> None:
  """Writes a string to a PDF file at the specified output path.

  Args:
      input_text (str): The text content to write to the PDF file
      output_path (str): The path where the PDF file will be saved
  Returns:
      None
  """
  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Create the PDF document
  doc = SimpleDocTemplate(output_path, pagesize=letter, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)

  # Get styles
  styles = getSampleStyleSheet()
  styles.add(ParagraphStyle(name='Custom', parent=styles['Normal'], fontSize=12, spaceAfter=12))

  # Create content
  content = []

  # Add timestamp
  content.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', styles['Custom']))
  content.append(Spacer(1, 0.25 * inch))

  # Add main content
  content.append(Paragraph(input_text, styles['Custom']))

  # Build the PDF
  doc.build(content)


def write_to_csv(data: List[List], output_path: str, headers: List[str] = None) -> str:
  """Writes data to a CSV file at the specified output path.

  Args:
      data (List[List]): A 2D list of data to write to the CSV file
      output_path (str): The path where the CSV file will be saved
      headers (List[str], optional): Column headers for the CSV file

  Returns:
      str: A success message with the output path
  """
  # Create directory if it doesn't exist
  os.makedirs(os.path.dirname(output_path), exist_ok=True)

  # Write the data to a CSV file
  with open(output_path, 'w', newline='') as f:
    writer = csv.writer(f)

    # Write headers if provided
    if headers:
      writer.writerow(headers)

    # Write data rows
    writer.writerows(data)

  return f'Successfully wrote data to CSV file at {output_path}'
